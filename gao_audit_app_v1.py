# gao_audit_app_v1.py
# Streamlit UI wrapper — GAO schedule audit with smart constraints + slack analysis (days)

import io
import re
from collections import Counter

import pandas as pd
import streamlit as st

# Try to enable circular dependency detection (optional)
try:
    import networkx as nx
    _HAS_NX = True
except Exception:
    _HAS_NX = False

DEFAULT_OUTNAME = "GAO_Schedule_Audit_Report_v4.xlsx"

# =========================
# Core audit helpers
# =========================
def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize column names and ensure required columns exist with safe defaults."""
    df = df.copy()
    df.columns = [c.strip().replace(" ", "_") for c in df.columns]
    rename_map = {
        "Unique_ID": "UID",
        "ID": "UID",
        "Start": "Start_Date",
        "Finish": "Finish_Date",
        "TotalSlack": "Total_Slack",
        "ConstraintType": "Constraint_Type",
        "Constraint_Date": "Constraint_Date",
        "Resource_Names": "ResourceNames",
    }
    df.rename(columns={c: rename_map.get(c, c) for c in df.columns}, inplace=True)

    if "UID" not in df.columns:
        df["UID"] = range(1, len(df) + 1)

    for col in ["Name", "Predecessors", "Successors", "Constraint_Type", "ResourceNames"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str)

    for b in ["Summary", "Milestone"]:
        if b not in df.columns:
            df[b] = False
        df[b] = df[b].fillna(False).astype(bool)

    # Dates
    for d in ["Start_Date", "Finish_Date", "Baseline_Start", "Baseline_Finish",
              "Actual_Start", "Actual_Finish", "Deadline", "Constraint_Date"]:
        if d in df.columns:
            df[d] = pd.to_datetime(df[d], errors="coerce")

    # Numeric
    for n in ["Total_Slack", "Duration", "Baseline_Duration"]:
        if n in df.columns:
            df[n] = pd.to_numeric(df[n], errors="coerce")

    return df


def parse_dependency_tokens(text: str):
    """Parse predecessor strings like '101FS, 102SS+3d, 103FF-2w'."""
    tokens, invalid = [], []
    if not isinstance(text, str) or not text.strip():
        return tokens, invalid
    for raw in re.split(r'[,\s]+', text.strip()):
        if not raw:
            continue
        m = re.match(r'^(\d+)([FS]{1,2})?([+-]?\d+[hdwmy])?$', raw, re.IGNORECASE)
        if m:
            pred = int(m.group(1))
            typ = (m.group(2) or 'FS').upper()
            off = m.group(3)
            tokens.append((pred, typ, off))
        else:
            invalid.append(raw)
    return tokens, invalid


def analyze_constraints(df: pd.DataFrame):
    """
    GAO-style constraint analysis:
    - Soft/valid (no penalty): ASAP/ALAP/None/blank
    - Hard/limiting (penalize): MSO, MFO, SNET, SNLT, FNET, FNLT
    Returns: (hard_count, soft_count, hard_detail_df)
    """
    ccol = next((c for c in df.columns if "constraint_type" in c.lower()), None)
    if not ccol:
        return 0, 0, pd.DataFrame(columns=["UID", "Name", "Constraint_Type"])

    s = df[ccol].fillna("").astype(str).str.strip().str.lower()
    soft = {"as soon as possible", "as late as possible", "none", ""}
    hard = {
        "must start on",
        "must finish on",
        "start no earlier than",
        "start no later than",
        "finish no earlier than",
        "finish no later than",
    }
    s_norm = s.replace({"asap": "as soon as possible", "alap": "as late as possible"})
    soft_mask = s_norm.isin(soft)
    hard_mask = s_norm.isin(hard)

    hard_rows = df[hard_mask].copy()
    detail_df = hard_rows[["UID", "Name", ccol]].rename(columns={ccol: "Constraint_Type"})
    return int(hard_mask.sum()), int(soft_mask.sum()), detail_df


def _slack_days(series: pd.Series) -> pd.Series:
    """
    Convert Total_Slack to days.
    Heuristic: if median absolute value > 200 → assume minutes (divide by 480), else already in days.
    """
    s = pd.to_numeric(series, errors="coerce")
    med = s.dropna().abs().median() if not s.dropna().empty else 0
    if med > 200:
        return s / 480.0
    return s


def run_audit_dataframe(df: pd.DataFrame, slack_lo_days: float, slack_hi_days: float, exclude_summaries: bool = True):
    df = _normalize_columns(df)
    df_main = df[df["Summary"] == False].copy() if exclude_summaries and "Summary" in df.columns else df.copy()

    # 1) Dependencies
    valid_links, invalid_rows = [], []
    for _, row in df_main.iterrows():
        tokens, invalid = parse_dependency_tokens(row.get("Predecessors", ""))
        if invalid:
            invalid_rows.append({"UID": row["UID"], "Name": row["Name"], "Malformed_Tokens": ", ".join(invalid)})
        for pred_id, typ, _ in tokens:
            valid_links.append((pred_id, row["UID"], typ))

    # 2) Circulars (optional)
    cycles = []
    if _HAS_NX and valid_links:
        G = nx.DiGraph()
        G.add_edges_from([(a, b) for a, b, _ in valid_links])
        try:
            cycles = list(nx.find_cycle(G, orientation="original"))
        except nx.NetworkXNoCycle:
            cycles = []

    dep_type_counts = Counter([t for _, _, t in valid_links])

    # 3) Dangling (no links)
    all_ids = set(df_main["UID"].tolist())
    connected = {a for a, _, _ in valid_links} | {b for _, b, _ in valid_links}
    dangling_ids = all_ids - connected
    dangling_tasks = df_main[df_main["UID"].isin(dangling_ids)][["UID", "Name"]]

    # 4) OOS actuals
    oos = pd.DataFrame(columns=["UID", "Name", "Detail"])
    if "Actual_Start" in df_main.columns and df_main["Actual_Start"].notna().any():
        for _, row in df_main.iterrows():
            astart = row.get("Actual_Start")
            if pd.isna(astart):
                continue
            tokens, _ = parse_dependency_tokens(row.get("Predecessors", ""))
            for pred_id, _, _ in tokens:
                pred_row = df_main[df_main["UID"] == pred_id]
                if not pred_row.empty:
                    pred_finish = pred_row.iloc[0].get("Finish_Date")
                    if pd.notna(pred_finish) and astart < pred_finish:
                        oos.loc[len(oos)] = [row["UID"], row["Name"], f"Started before pred {pred_id} finished"]
                        break

    # 5) Baseline variance (late count)
    baseline_variance = pd.DataFrame(columns=["UID", "Name", "Finish_Date", "Baseline_Finish", "Variance_Days"])
    late_count = 0
    if "Baseline_Finish" in df_main.columns and "Finish_Date" in df_main.columns:
        with_base = df_main.dropna(subset=["Baseline_Finish"]).copy()
        if not with_base.empty:
            vd = (with_base["Finish_Date"] - with_base["Baseline_Finish"]).dt.days
            baseline_variance = with_base.assign(Variance_Days=vd)[["UID", "Name", "Finish_Date", "Baseline_Finish", "Variance_Days"]]
            late_count = (baseline_variance["Variance_Days"].fillna(0) > 0).sum()

    # 6) Constraints (hard vs soft)
    hard_count, soft_count, hard_detail = analyze_constraints(df_main)

    # 7) Slack analysis (only flagged rows)
    slack_days = _slack_days(df_main.get("Total_Slack", pd.Series(dtype=float)))
    slack_flags = []
    for idx in df_main.index:
        val = slack_days.loc[idx] if idx in slack_days.index else None
        if pd.isna(val):
            continue
        if val < float(slack_lo_days):
            slack_flags.append({"UID": df_main.at[idx, "UID"], "Name": df_main.at[idx, "Name"],
                                "Total_Slack_Days": float(val), "Slack_Type": "Negative Slack"})
        elif val > float(slack_hi_days):
            slack_flags.append({"UID": df_main.at[idx, "UID"], "Name": df_main.at[idx, "Name"],
                                "Total_Slack_Days": float(val), "Slack_Type": "Excessive Slack"})

    slack_analysis = pd.DataFrame(slack_flags)

    summary_items = {
        "Malformed/Missing Links": len(invalid_rows),
        "Circular Dependencies": len(cycles),
        "Dangling Tasks": len(dangling_tasks),
        "Out-of-Sequence Actuals": len(oos),
        "Baseline Variance (Late)": int(late_count),
        "Constraints (Hard)": int(hard_count),
        "Constraints (Soft / Valid)": int(soft_count),
        "Negative Slack": int((slack_analysis["Slack_Type"] == "Negative Slack").sum()),
        "Excessive Slack": int((slack_analysis["Slack_Type"] == "Excessive Slack").sum()),
    }
    summary_df = pd.DataFrame(list(summary_items.items()), columns=["Metric", "Value"])

    # Meta
    meta = pd.DataFrame([{"Tasks_NonSummary": int(len(df_main))}])

    # Return structured audit dict
    return {
        "summary": summary_df,
        "invalid_rows": pd.DataFrame(invalid_rows),
        "cycles": pd.DataFrame(cycles, columns=["From", "To", "Type"]) if cycles else pd.DataFrame(columns=["From","To","Type"]),
        "dangling": dangling_tasks,
        "oos": oos,
        "baseline_variance": baseline_variance,
        "constraints_hard": hard_detail,
        "dep_type_counts": pd.DataFrame(list(dep_type_counts.items()), columns=["Dependency_Type","Count"]),
        "slack_analysis": slack_analysis,  # only flagged rows
        "meta": meta,
    }


def build_report_bytes(audit_dict, output_name=DEFAULT_OUTNAME) -> bytes:
    """Create the Excel report in-memory and return bytes."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        for name, df in audit_dict.items():
            sname = name[:31]
            df.to_excel(w, sheet_name=sname, index=False)

    bio.seek(0)
    wb = load_workbook(bio)

    # Scoring
    ws = wb["summary"]
    # Weights (soft constraints not penalized)
    weights = {
        "Malformed/Missing Links": 0.5,
        "Circular Dependencies": 2.0,
        "Dangling Tasks": 0.2,
        "Out-of-Sequence Actuals": 0.8,
        "Baseline Variance (Late)": 0.6,
        "Constraints (Hard)": 0.5,
        "Negative Slack": 1.0,
        "Excessive Slack": 0.4,
    }
    tasks_count = 1
    if "meta" in wb.sheetnames:
        mws = wb["meta"]
        try:
            # find Tasks_NonSummary in col A, value in col B
            for r in range(2, mws.max_row + 1):
                if str(mws.cell(r, 1).value).strip() == "Tasks_NonSummary":
                    val = mws.cell(r, 2).value
                    if isinstance(val, (int, float)) and val:
                        tasks_count = max(1, int(val))
                    break
        except Exception:
            pass

    # Compute penalty
    penalty = 0.0
    # locate summary header (Metric, Value)
    # assuming headers at row 1
    for r in range(2, ws.max_row + 1):
        metric = ws.cell(r, 1).value
        value = ws.cell(r, 2).value
        if metric in (None, "Schedule Integrity Score (0–100)", "Project Health"):
            continue
        if metric == "Constraints (Soft / Valid)":
            continue
        w = weights.get(metric, 0.2)
        if metric in {"Constraints (Hard)", "Dangling Tasks"}:
            try:
                value = min(int(value), 1000)
            except Exception:
                value = 0
        try:
            penalty += float(value) * w
        except Exception:
            pass

    score = max(0.0, 100.0 - (penalty / max(tasks_count, 1) * 100.0))
    score = round(score, 1)

    ws.append(["Schedule Integrity Score (0–100)", score])
    status = "Excellent ✅" if score > 90 else "Good 🟡" if score > 75 else "Needs Work 🔴"
    ws.append(["Project Health", status])

    # Style header
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="DDDDDD")

    # Heatmap (exclude last 2 rows)
    last_two = 2
    for r in range(2, max(2, ws.max_row - last_two + 1)):
        val = ws.cell(r, 2).value
        color = "92D050"
        if isinstance(val, (int, float)):
            if val > 200:
                color = "FF6666"
            elif val > 20:
                color = "FFD966"
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=color)

    # Pie chart for dependency types
    if "dep_type_counts" in wb.sheetnames:
        ct = wb["dep_type_counts"]
        if ct.max_row >= 2:
            chart = PieChart()
            chart.title = "Dependency Type Distribution"
            labels = Reference(ct, min_col=1, min_row=2, max_row=ct.max_row)
            data = Reference(ct, min_col=2, min_row=1, max_row=ct.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 7
            chart.width = 8
            chart.dLbls = DataLabelList()
            chart.dLbls.showVal = True
            chart.dLbls.showPercent = True
            ws.add_chart(chart, "G4")

    # Auto-fit summary columns
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    # Write back to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), score, status


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="GAO Schedule Quality Auditor", layout="wide", page_icon="📘")
st.title("GAO Schedule Quality Auditor")
st.caption("Automated Schedule Health & Logic Analysis — GAO-inspired checks")

with st.sidebar:
    st.header("⚙️ Settings")
    slack_lo = st.number_input("Slack Low (days) → Negative if below", value=0.0, step=1.0)
    slack_hi = st.number_input("Slack High (days) → Excessive if above", value=60.0, step=5.0)
    exclude_summaries = st.checkbox("Exclude Summary tasks from logic checks", value=True)
    out_name = st.text_input("Output file name", value=DEFAULT_OUTNAME)

st.markdown("### 📤 Upload your Microsoft Project Export (.xlsx)")
uploaded = st.file_uploader("Drag & drop or browse", type=["xlsx"])

if uploaded is not None:
    with st.spinner("Analyzing schedule…"):
        # Read first sheet by default (or let user choose later)
        xls = pd.ExcelFile(uploaded)
        sheet_to_use = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_to_use)

        audit = run_audit_dataframe(df, slack_lo_days=slack_lo, slack_hi_days=slack_hi, exclude_summaries=exclude_summaries)
        report_bytes, score, status = build_report_bytes(audit, out_name)

    st.success(f"✅ Analysis complete — Score: {score}/100 | {status}")
    # KPI row
    cols = st.columns(3)
    cols[0].metric("Schedule Integrity Score", f"{score}/100")
    cols[1].metric("Project Health", status)
    cols[2].metric("Tasks (non-summary)", int(audit["meta"]["Tasks_NonSummary"].iloc[0]) if not audit["meta"].empty else 0)

    # Tabs
    tabs = st.tabs(["Summary", "Slack Analysis", "Constraints (Hard)", "Baseline Variance", "Dependencies", "Malformed Links"])
    with tabs[0]:
        st.subheader("Summary")
        st.dataframe(audit["summary"], use_container_width=True)
        st.download_button(
            "⬇️ Download Excel Report",
            data=report_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with tabs[1]:
        st.subheader("Slack Analysis (flagged only)")
        st.caption(f"Negative if < {slack_lo} days; Excessive if > {slack_hi} days")
        st.dataframe(audit["slack_analysis"], use_container_width=True, height=400)

    with tabs[2]:
        st.subheader("Hard Constraints (penalized)")
        st.dataframe(audit["constraints_hard"], use_container_width=True, height=400)

    with tabs[3]:
        st.subheader("Baseline Variance")
        st.dataframe(audit["baseline_variance"], use_container_width=True, height=400)

    with tabs[4]:
        st.subheader("Dependencies")
        st.dataframe(audit["dep_type_counts"], use_container_width=True, height=300)

        st.subheader("Dangling Tasks")
        st.dataframe(audit["dangling"], use_container_width=True, height=300)

        if _HAS_NX:
            st.subheader("Circular Dependencies")
            st.dataframe(audit["cycles"], use_container_width=True, height=300)
        else:
            st.info("Install `networkx` to enable circular dependency detection.")

    with tabs[5]:
        st.subheader("Malformed / Missing Predecessors")
        st.dataframe(audit["invalid_rows"], use_container_width=True, height=400)

else:
    st.info("Upload a Microsoft Project Excel export (.xlsx) to begin.")
