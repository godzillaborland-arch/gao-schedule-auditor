import os, re, math, pandas as pd

# --- Config ---
LEAD_LAG_ABS_LIMIT = 10          # flag |lead/lag| > 10 (d/w/m/y offsets in dependency text)
LONG_DURATION_DAYS = 60          # >60 working days = long
SHORT_DURATION_MIN = 60          # <60 minutes = short
VALID_CONSTRAINTS = {
    # treat only these as real constraints (ignore ASAP/NA/defaults)
    "must start on","must finish on",
    "start no earlier than","start no later than",
    "finish no earlier than","finish no later than"
}
WEIGHTS = {
    "Malformed/Missing Links": 2.0,
    "Circular Dependencies": 3.5,
    "Dangling Tasks": 2.0,
    "Lead/Lag Warnings": 1.5,
    "Critical Path Logic Gaps": 1.5,
    "Negative Slack": 2.0,
    "Excessive Slack": 1.0,
    "Constraints": 1.0,
    "No Resources": 1.0,
    "Unrealistic Duration": 1.0,
    "Isolated Milestone": 1.0,
    "Low-Quality Name": 0.5,
    "Tasks Behind Baseline": 1.5,   # baseline penalty
}

# --- Helpers ---
def _only_digits(text, default=None):
    if text is None: return default
    m = re.search(r"-?\d+\.?\d*", str(text))
    return float(m.group()) if m else default

def minutes_to_days(minutes):
    try: return float(minutes)/480.0
    except: return None

def duration_flags(minutes):
    flags = []
    val = _only_digits(minutes)
    if val is None:
        return flags
    d = minutes_to_days(val)
    if d is not None and d > LONG_DURATION_DAYS:
        flags.append("Unrealistic Duration: Long")
    if val < SHORT_DURATION_MIN:
        flags.append("Unrealistic Duration: Short")
    return flags

def parse_dependency_tokens(text):
    """Accepts '101FS, 102SS+3d, 103FF-2w' → [(pred:int, type:str, offset:str|None)], [invalid]"""
    toks, invalid = [], []
    if not isinstance(text, str) or not text.strip():
        return toks, invalid
    for raw in re.split(r"[,\s]+", text.strip()):
        if not raw:
            continue
        m = re.match(r"^(\d+)([FS]{1,2})?([+-]?\d+[hdwmy])?$", raw, re.I)
        if m:
            pred = int(m.group(1))
            typ  = (m.group(2) or "FS").upper()
            off  = m.group(3)
            toks.append((pred, typ, off))
        else:
            invalid.append(raw)
    return toks, invalid

def find_cycles_directed(edges):
    """Simple DFS cycle counter for (u,v) edges; returns count of back-edges (proxy for cycles)."""
    from collections import defaultdict
    g = defaultdict(list)
    for u,v in edges: g[u].append(v)
    seen, onstack = set(), set()
    cycles = 0
    def dfs(u):
        nonlocal cycles
        seen.add(u); onstack.add(u)
        for v in g[u]:
            if v not in seen:
                dfs(v)
            elif v in onstack:
                cycles += 1
        onstack.remove(u)
    for n in list(g.keys()):
        if n not in seen:
            dfs(n)
    return cycles

def to_bool(x):
    s = str(x).strip().lower()
    return s in ("true","t","1","yes","y","x")

def parse_datetime(s):
    return pd.to_datetime(s, errors="coerce")

# --- Load Excel ---
def load_schedule(path, sheet_name="Task_Table1"):
    try:
        df = pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        first = pd.ExcelFile(path).sheet_names[0]
        df = pd.read_excel(path, sheet_name=first)

    # rename common variants → canonical names
    rename_map = {}
    for col in list(df.columns):
        low = col.strip().lower()
        if low in ("unique_id","uniqueid","unique i_d","uid"): rename_map[col] = "Unique_ID"
        elif low in ("id",): rename_map[col] = "ID"
        elif low in ("name","task name"): rename_map[col] = "Name"
        elif low in ("start","start_date","start date"): rename_map[col] = "Start_Date"
        elif low in ("finish","finish_date","finish date"): rename_map[col] = "Finish_Date"
        elif low in ("predecessors",): rename_map[col] = "Predecessors"
        elif low in ("successors",): rename_map[col] = "Successors"
        elif low in ("resource names","resourcenames"): rename_map[col] = "ResourceNames"
        elif low in ("total slack","total_slack","totalslack"): rename_map[col] = "Total_Slack"
        elif low in ("constraint type","constraint_type","constrainttype"): rename_map[col] = "ConstraintType"
        elif low in ("baseline start","baseline_start"): rename_map[col] = "Baseline_Start"
        elif low in ("baseline finish","baseline_finish"): rename_map[col] = "Baseline_Finish"
        elif low in ("milestone",): rename_map[col] = "Milestone"
        elif low in ("summary",): rename_map[col] = "Summary"
        elif low in ("duration",): rename_map[col] = "Duration"
        elif low in ("actual start","actual_start"): rename_map[col] = "Actual_Start"
        elif low in ("actual finish","actual_finish"): rename_map[col] = "Actual_Finish"
    if rename_map:
        df = df.rename(columns=rename_map)

    # ensure expected columns
    for c in ["Unique_ID","ID","Name","Start_Date","Finish_Date","Predecessors","Successors",
              "ResourceNames","Total_Slack","ConstraintType","Baseline_Start","Baseline_Finish",
              "Milestone","Summary","Duration","Actual_Start","Actual_Finish"]:
        if c not in df.columns:
            df[c] = None

    # UID
    if pd.notna(df["Unique_ID"]).any():
        df["UID"] = df["Unique_ID"]
    elif pd.notna(df["ID"]).any():
        df["UID"] = df["ID"]
    else:
        df["UID"] = range(1, len(df)+1)

    # text norm
    for c in ["Name","Predecessors","Successors","ResourceNames","ConstraintType"]:
        df[c] = df[c].astype(str).replace("nan","").fillna("")

    # booleans
    df["Milestone"] = df["Milestone"].apply(to_bool)
    df["Summary"]   = df["Summary"].apply(to_bool)

    # numerics
    df["Total_Slack"] = df["Total_Slack"].apply(_only_digits)
    df["Duration"]    = df["Duration"].apply(_only_digits)

    # dates
    for dc in ["Start_Date","Finish_Date","Baseline_Start","Baseline_Finish","Actual_Start","Actual_Finish"]:
        df[dc+"_dt"] = parse_datetime(df[dc])

    # name fallback
    df["Name"] = df["Name"].replace("", "(unnamed)")
    return df

# --- Audit ---
def run_audit(df):
    uid = df["UID"].tolist()
    name = df["Name"].astype(str).tolist()
    preds = df["Predecessors"].astype(str).tolist()
    succs = df["Successors"].astype(str).tolist()
    res   = df["ResourceNames"].astype(str).tolist()
    is_sum = [bool(x) for x in df["Summary"].tolist()]
    is_ms  = [bool(x) for x in df["Milestone"].tolist()]

    # 1) Parse dependencies
    valid_links, invalid_rows, lead_lag = [], [], []
    dep_type_counts = {}
    for i in range(len(df)):
        tokens, bad = parse_dependency_tokens(preds[i])
        if bad:
            invalid_rows.append({"UID": uid[i], "Name": name[i], "Malformed": ", ".join(bad)})
        for pid, typ, off in tokens:
            if off:
                try:
                    n = int(re.findall(r"[+-]?\d+", off)[0])
                    if abs(n) > LEAD_LAG_ABS_LIMIT:
                        lead_lag.append({"UID": uid[i], "Name": name[i], "LeadLag": off})
                except:
                    pass
            valid_links.append((pid, uid[i], typ))
            dep_type_counts[typ] = dep_type_counts.get(typ, 0) + 1

    # 2) Circulars (proxy via DFS back-edges)
    edges = [(a,b) for (a,b,_) in valid_links]
    cycles_count = find_cycles_directed(edges)

    # 3) Connectivity & FS logic
    all_ids = set(uid)
    connected = set([a for a,_,_ in valid_links]) | set([b for _,b,_ in valid_links])
    dangling_ids = all_ids - connected
    dangling = [{"UID": u, "Name": df.loc[df["UID"]==u, "Name"].iloc[0]} for u in sorted(dangling_ids)]
    incoming_fs = set([b for a,b,t in valid_links if t=="FS"])
    outgoing_fs = set([a for a,b,t in valid_links if t=="FS"])
    no_fs = all_ids - (incoming_fs | outgoing_fs)
    cp_gaps = [{"UID": u, "Name": df.loc[df["UID"]==u, "Name"].iloc[0]} for u in sorted(no_fs)]

    # 4) Slack, constraints, resources, durations, milestones, names
    neg_slack, big_slack, constraints, no_res, unreal, iso_ms, bad_names = [], [], [], [], [], [], []
    for i in range(len(df)):
        ts = df.iloc[i]["Total_Slack"]
        if ts is not None and not (isinstance(ts, float) and math.isnan(ts)):
            if ts < 0:
                neg_slack.append({"UID": uid[i], "Name": name[i], "TotalSlack": ts})
            if ts > 55000:
                big_slack.append({"UID": uid[i], "Name": name[i], "TotalSlack": ts})

        cstr = str(df.iloc[i]["ConstraintType"] or "").strip().lower()
        if cstr in VALID_CONSTRAINTS:
            constraints.append({"UID": uid[i], "Name": name[i], "ConstraintType": cstr})

        if (not is_sum[i]) and (res[i].strip() == ""):
            no_res.append({"UID": uid[i], "Name": name[i]})

        for f in duration_flags(df.iloc[i]["Duration"]):
            unreal.append({"UID": uid[i], "Name": name[i], "Flag": f})

        if is_ms[i] and (str(preds[i]).strip()=="" or str(succs[i]).strip()==""):
            iso_ms.append({"UID": uid[i], "Name": name[i]})

        if (not is_sum[i]) and len(name[i]) < 10:
            bad_names.append({"UID": uid[i], "Name": name[i]})

    # 5) Baseline variance (Finish vs Baseline Finish)
    baseline_variance_rows = []
    for i in range(len(df)):
        f = df.iloc[i]["Finish_Date_dt"]
        b = df.iloc[i]["Baseline_Finish_dt"]
        if pd.notna(f) and pd.notna(b):
            var = (f - b).days
            if isinstance(var, (int, float)) and var > 0:
                baseline_variance_rows.append({"UID": uid[i], "Name": name[i], "VarianceDays": int(var)})
    top10 = sorted(baseline_variance_rows, key=lambda r: r["VarianceDays"], reverse=True)[:10]

    # Summary counts
    summary_items = {
        "Malformed/Missing Links": len(invalid_rows),
        "Circular Dependencies": cycles_count,
        "Dangling Tasks": len(dangling),
        "Lead/Lag Warnings": len(lead_lag),
        "Critical Path Logic Gaps": len(cp_gaps),
        "Negative Slack": len(neg_slack),
        "Excessive Slack": len(big_slack),
        "Constraints": len(constraints),
        "No Resources": len(no_res),
        "Unrealistic Duration": len(unreal),
        "Isolated Milestone": len(iso_ms),
        "Low-Quality Name": len(bad_names),
        "Tasks Behind Baseline": len(baseline_variance_rows),
    }
    summary_df = pd.DataFrame(list(summary_items.items()), columns=["Metric","Value"])
    dep_types_df = pd.DataFrame([{"Dependency_Type": k, "Count": v} for k,v in dep_type_counts.items()])

    # Score
    num_tasks = max(1, len(df))
    penalty = sum(summary_items[m]*WEIGHTS.get(m,0.0) for m in summary_items)
    score = max(0.0, 100.0 - (penalty / num_tasks) * 100.0)
    score = round(score, 1)
    health = "Excellent 🟢" if score >= 90 else "Good 🟡" if score >= 75 else "Needs Attention 🔴"

    return {
        "Summary": summary_df,
        "Dependency_Types": dep_types_df,
        "Malformed_Missing": pd.DataFrame(invalid_rows),
        "Circular_Dependencies": pd.DataFrame({"Cycles_Found":[cycles_count]}),
        "Dangling_Tasks": pd.DataFrame(dangling),
        "Lead_Lag_Warnings": pd.DataFrame(lead_lag),
        "Critical_Path_Issues": pd.DataFrame(cp_gaps),
        "Negative_Slack": pd.DataFrame(neg_slack),
        "Excessive_Slack": pd.DataFrame(big_slack),
        "Constraints": pd.DataFrame(constraints),
        "No_Resources": pd.DataFrame(no_res),
        "Unrealistic_Duration": pd.DataFrame(unreal),
        "Isolated_Milestones": pd.DataFrame(iso_ms),
        "Low_Quality_Names": pd.DataFrame(bad_names),
        "Baseline_Variance": pd.DataFrame(baseline_variance_rows),
        "Top10_Delayed": pd.DataFrame(top10),
        "_score": score,
        "_health": health,
    }

# --- Write Excel ---
def write_report(audit, outfile):
    with pd.ExcelWriter(outfile, engine="openpyxl") as w:
        for name, dfpart in audit.items():
            if name.startswith("_"):  # internal keys
                continue
            (dfpart if isinstance(dfpart, pd.DataFrame) else pd.DataFrame(dfpart)).to_excel(
                w, sheet_name=name[:31], index=False
            )

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = load_workbook(outfile)
    ws = wb["Summary"]

    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="Schedule Health Score (0–100)")
    ws.cell(row=last, column=2, value=audit["_score"])
    ws.cell(row=last+1, column=1, value="Project Health")
    ws.cell(row=last+1, column=2, value=audit["_health"])

    # header style
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # traffic light
    score = audit["_score"]
    score_cell = ws.cell(row=last, column=2)
    if score >= 90:
        score_cell.fill = PatternFill("solid", fgColor="92D050")
    elif score >= 75:
        score_cell.fill = PatternFill("solid", fgColor="FFD966")
    else:
        score_cell.fill = PatternFill("solid", fgColor="FF6666")

    wb.save(outfile)
    return audit["_score"], audit["_health"]
